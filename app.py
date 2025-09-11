from flask import Flask, render_template, request, send_file
from simple_salesforce import Salesforce, SalesforceAuthenticationFailed, SalesforceMalformedRequest
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import csv
import random
from werkzeug.utils import secure_filename
import io
import zipfile

app = Flask(__name__)

sf_connection = None
empty_columns_results_storage = []
duplicate_columns_results_storage = []

# Setup for file uploads
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        security_token = request.form['token']
        domain = request.form.get('environment', 'login')

        try:
            global sf_connection
            sf_connection = Salesforce(username=username, password=password, security_token=security_token, domain=domain)
            all_objects = sf_connection.describe()['sobjects']
            return render_template('select_objects.html', all_objects=all_objects)
        except Exception as e:
            return f"Failed to authenticate: {str(e)}"
    else:
        return render_template('index.html')

@app.route('/export_records_login', methods=['GET', 'POST'])
def export_records_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        security_token = request.form['token']
        domain = request.form.get('environment', 'login')
        try:
            global sf_connection
            sf_connection = Salesforce(username=username, password=password, security_token=security_token, domain=domain)
            all_objects = sf_connection.describe()['sobjects']
            return render_template('select_records.html', all_objects=all_objects)
        except Exception as e:
            return f"Failed to authenticate: {str(e)}"
    else:
        return render_template('records_login.html')

@app.route('/export', methods=['POST'])
def export_objects():
    selected_objects = request.form.getlist('objects')
    text_to_remove = request.form.get('text_remove')
    column_name = request.form.get('column_name', '2GP Fields')

    uploaded_file = request.files.get('file')
    if uploaded_file and uploaded_file.filename != '':
        uploaded_excel = pd.ExcelFile(uploaded_file)
    else:
        uploaded_excel = None

    objects_df = pd.DataFrame(selected_objects, columns=['Object Name'])
    excel_file = 'salesforce_objects.xlsx'
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        objects_df.to_excel(writer, sheet_name='Objects List', index=False)

        for obj_name in selected_objects:
            fields = getattr(sf_connection, obj_name).describe()['fields']
            fields_df = pd.DataFrame(fields)

            headers_df = pd.DataFrame(columns=['S#', '1GP Fields', '2GP Fields', 'Field API Name', 'Field Type', f'Object Name ({obj_name})'])

            fields_data = pd.DataFrame({
                'S#': range(1, len(fields_df) + 1),
                '1GP Fields': fields_df['name'],
                '2GP Fields': [None] * len(fields_df),
                'Field API Name': fields_df['name'],
                'Field Type': fields_df['type'],
                f'Object Name ({obj_name})': [None] * len(fields_df)
            })

            if text_to_remove:
                fields_data['1GP Fields'] = fields_data['1GP Fields'].str.replace(text_to_remove, '')

            final_df = pd.concat([headers_df, fields_data], ignore_index=True)

            if uploaded_excel and obj_name in uploaded_excel.sheet_names:
                uploaded_df = pd.read_excel(uploaded_excel, sheet_name=obj_name)
                df22gp = set(uploaded_df.loc[:, column_name])
                df11gp = set(final_df.loc[:, "1GP Fields"])
                df2e = df22gp.difference(df11gp)
                df1e = df11gp.difference(df22gp)

                if df2e == df1e:
                    print("Same")
                else:
                    for i in range(len(final_df)):
                        if final_df.at[i, '1GP Fields'] in df1e:
                            pass
                        else:
                            final_df.loc[i, ['2GP Fields']] = [final_df.at[i, '1GP Fields']]
                    for i in df2e:
                        new_row = {'S#': None, '1GP Fields': None, '2GP Fields': i, 'Field Type': None}
                        final_df = pd.concat([final_df, pd.DataFrame([new_row])], ignore_index=True)

            sheet_name = obj_name[:31] if len(obj_name) > 31 else obj_name
            final_df.to_excel(writer, sheet_name=sheet_name, index=False)

    return send_file(excel_file, as_attachment=True, download_name='salesforce_objects.xlsx')


@app.route('/export_records', methods=['POST'])
def export_records():
    selected_objects = request.form.getlist('objects')
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for obj_name in selected_objects:
            fields_info = getattr(sf_connection, obj_name).describe()['fields']
            field_names = [f['name'] for f in fields_info]
            query = f"SELECT {', '.join(field_names)} FROM {obj_name}"
            records = sf_connection.query_all(query)
            df = pd.DataFrame(records['records'])
            if 'attributes' in df.columns:
                df = df.drop(columns='attributes')
            csv_data = df.to_csv(index=False)
            zf.writestr(f"{obj_name}.csv", csv_data)
    zip_buffer.seek(0)
    return send_file(zip_buffer, as_attachment=True, download_name='object_records.zip', mimetype='application/zip')


@app.route('/map_fields', methods=['GET', 'POST'])
def map_fields():
    if request.method == 'POST':
        file1 = request.files.get('file1')
        file2 = request.files.get('file2')
        row1_name = request.form.get('row1')
        row2_name = request.form.get('row2')

        if file1 and file2:
            excel_file1 = pd.ExcelFile(file1)
            excel_file2 = pd.ExcelFile(file2)

            output_file = 'updated_file.xlsx'
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for sheet_name in excel_file1.sheet_names:
                    df1 = pd.read_excel(excel_file1, sheet_name=sheet_name)

                    # Check if the sheet also exists in file2
                    if sheet_name in excel_file2.sheet_names:
                        df2 = pd.read_excel(excel_file2, sheet_name=sheet_name)
                        
                        try:
                            df22gp = set(df2.loc[:, row2_name].dropna())
                        except KeyError:
                            print(f"Column name '{row2_name}' not found in sheet '{sheet_name}' of file 2. Skipping to the next sheet.")
                            df1.to_excel(writer, sheet_name=sheet_name, index=False)
                            continue

                        df11gp = set(df1.loc[:, row1_name].dropna())
                        df2e = df22gp.difference(df11gp)
                        df1e = df11gp.difference(df22gp)

                        matched = df11gp.intersection(df22gp)

                        # New logic to remove unmatched data from row1 and prepare for appending unmatched data
                        df1 = df1[df1[row1_name].isin(matched) | df1[row1_name].isnull()]
                        unmatched_row1_df = pd.DataFrame({row1_name: list(df1e), row2_name: [None]*len(df1e)})
                        unmatched_row2_df = pd.DataFrame({row1_name: [None]*len(df2e), row2_name: list(df2e)})

                        for i in df1.index:
                            if df1.at[i, row1_name] in matched:
                                df1.at[i, row2_name] = df1.at[i, row1_name]

                        df1 = pd.concat([df1, unmatched_row1_df, unmatched_row2_df], ignore_index=True)
                        df1.to_excel(writer, sheet_name=sheet_name, index=False)

                        # Apply cell styling for matched fields
                        workbook = writer.book
                        worksheet = writer.sheets[sheet_name]
                        fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                        for i, value in enumerate(df1[row1_name], start=1):
                            if value in matched:
                                cell = worksheet.cell(row=i+1, column=df1.columns.get_loc(row1_name)+1)
                                cell.fill = fill
                                cell = worksheet.cell(row=i+1, column=df1.columns.get_loc(row2_name)+1)
                                cell.fill = fill
                    else:
                        # If the sheet name doesn't exist in file2, write df1 as is
                        df1.to_excel(writer, sheet_name=sheet_name, index=False)

            return send_file(output_file, as_attachment=True, download_name='updated_file.xlsx')
        else:
            return "Please upload both files."
    else:
        return render_template('map_fields.html')


def find_empty_columns_in_csv(folder_path):
    csv.field_size_limit(2147483647)  # Setting a large field size limit
    results = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            file_path = os.path.join(folder_path, filename)
            try:
                csvfile = open(file_path, newline='', encoding='utf-8')
            except UnicodeDecodeError:
                csvfile = open(file_path, newline='', encoding='latin-1')
            with csvfile:
                reader = csv.reader(csvfile)
                header = next(reader, None)  # Read the header row
                if header is None:
                    results.append({'filename': filename, 'empty_columns': []})
                    continue

                # Initialize a dictionary to track if a column is empty
                empty_columns = {column_name: True for column_name in header}

                for row in reader:
                    for column_name, cell in zip(header, row):
                        if cell.strip() != '':
                            empty_columns[column_name] = False

                # Filter out columns that are not empty
                empty_columns = [column_name for column_name, is_empty in empty_columns.items() if is_empty]

                results.append({'filename': filename, 'empty_columns': empty_columns})
    return results


def find_duplicate_columns_in_csv(folder_path):
    results = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            file_path = os.path.join(folder_path, filename)
            try:
                df = pd.read_csv(file_path, dtype=str)
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, dtype=str, encoding='latin-1')

            duplicates = []
            columns = list(df.columns)
            for i in range(len(columns)):
                for j in range(i + 1, len(columns)):
                    col1 = df[columns[i]].fillna('')
                    col2 = df[columns[j]].fillna('')
                    if col1.equals(col2):
                        duplicates.append((columns[i], columns[j]))

            results.append({'filename': filename, 'duplicate_columns': duplicates})
    return results
    
@app.route('/find_empty_columns', methods=['GET', 'POST'])
def find_empty_columns():
    global empty_columns_results_storage
    if request.method == 'POST':
        folder_path = request.form['folder_path']
        try:
            results = find_empty_columns_in_csv(folder_path)
            empty_columns_results_storage = results
            return render_template('empty_columns_results.html', results=results)
        except Exception as e:
            return f"An error occurred: {str(e)}"
    else:
        return render_template('empty_columns_form.html')

@app.route('/download_empty_columns_csv')
def download_empty_columns_csv():
    if not empty_columns_results_storage:
        return "No results to download."
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for result in empty_columns_results_storage:
            df = pd.DataFrame({'Empty Columns': result['empty_columns']})
            csv_data = df.to_csv(index=False)
            zf.writestr(f"{result['filename']}_empty_columns.csv", csv_data)
    memory_file.seek(0)
    return send_file(memory_file, as_attachment=True, download_name='empty_columns.zip', mimetype='application/zip')


@app.route('/find_duplicate_columns', methods=['GET', 'POST'])
def find_duplicate_columns():
    global duplicate_columns_results_storage
    if request.method == 'POST':
        folder_path = request.form['folder_path']
        try:
            results = find_duplicate_columns_in_csv(folder_path)
            duplicate_columns_results_storage = results
            return render_template('duplicate_columns_results.html', results=results)
        except Exception as e:
            return f"An error occurred: {str(e)}"
    else:
        return render_template('duplicate_columns_form.html')


@app.route('/download_duplicate_columns_csv')
def download_duplicate_columns_csv():
    if not duplicate_columns_results_storage:
        return "No results to download."
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for result in duplicate_columns_results_storage:
            df = pd.DataFrame(result['duplicate_columns'], columns=['Column 1', 'Column 2'])
            csv_data = df.to_csv(index=False)
            zf.writestr(f"{result['filename']}_duplicate_columns.csv", csv_data)
    memory_file.seek(0)
    return send_file(memory_file, as_attachment=True, download_name='duplicate_columns.zip', mimetype='application/zip')

@app.route('/salesforce_login', methods=['GET', 'POST'])
def salesforce_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        token = request.form['token']
        domain = request.form.get('environment', 'login')
        try:
            sf = Salesforce(username=username, password=password, security_token=token, domain=domain)
            object_names = [obj["name"] for obj in sf.describe()["sobjects"]]
            
            objects_with_records = 0
            errors = []  # To track errors for objects that cannot be queried
            for obj_name in object_names:
                try:
                    query = f"SELECT Id FROM {obj_name} LIMIT 1"
                    records = sf.query_all(query)
                    if records['totalSize'] > 0:
                        objects_with_records += 1
                except SalesforceMalformedRequest as e:
                    errors.append(f"Cannot query object {obj_name}: {e}")
                    # You can also log these errors or handle them as needed

            # Assuming you want to show the count and possibly errors in your template
            return render_template('salesforce_analysis_result.html', objects_with_records=objects_with_records, errors=errors)
        except SalesforceAuthenticationFailed as e:
            return f"Failed to authenticate with Salesforce: {e}"
    else:
        return render_template('salesforce_login.html')
    
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def load_data(file_path):
    """Load data from CSV or XLSX based on file extension."""
    if file_path.endswith('.csv'):
        return pd.read_csv(file_path).to_dict(orient='records')
    elif file_path.endswith('.xlsx'):
        return pd.read_excel(file_path).to_dict(orient='records')

def record_count_check(data1, data2):
    """Check if the record count matches between two datasets."""
    count1, count2 = len(data1), len(data2)
    if count1 == count2:
        return f"Verified. Total record count for both files: {count1}"
    return f"Record Count Mismatch. Count for File 1: {count1}, Count for File 2: {count2}"

def check_random_record(data1, data2, column_name1, column_name2):
    """Perform a random record check between two datasets."""
    if not data1 or not data2:
        return "One of the files is empty."

    random_record = random.choice(data1)
    if column_name1 not in random_record:
        return f"Error: Column '{column_name1}' not found in File 1."

    value_to_find = random_record[column_name1]
    matching_records = [row for row in data2 if row.get(column_name2) == value_to_find]

    if not matching_records:
        return "Random record matching failed."
    
    # Build matched and unmatched column data
    matched_columns, unmatched_columns = [], []
    for key, value in random_record.items():
        if key in matching_records[0] and value == matching_records[0][key]:
            matched_columns.append(key)
        else:
            unmatched_columns.append(key)
    return {"matched_columns": matched_columns, "unmatched_columns": unmatched_columns}

def compare_column_records(data1, data2):
    """Compare column records between two datasets."""
    columns1 = set(data1[0].keys()) if data1 else set()
    columns2 = set(data2[0].keys()) if data2 else set()
    matched_columns = columns1.intersection(columns2)
    unmatched_columns = columns1.symmetric_difference(columns2)
    return list(matched_columns), list(unmatched_columns)

@app.route('/validate_data', methods=['GET', 'POST'])
def validate_data():
    if request.method == 'POST':
        file1 = request.files['file1']
        file2 = request.files['file2']
        column_name1 = request.form.get('column_name1', 'ID')  # Default column name
        column_name2 = request.form.get('column_name2', 'PRODUCTION ID')

        if file1 and file2 and allowed_file(file1.filename) and allowed_file(file2.filename):
            # Temporarily save files for processing
            filename1 = secure_filename(file1.filename)
            file_path1 = os.path.join(app.config['UPLOAD_FOLDER'], filename1)
            file1.save(file_path1)

            filename2 = secure_filename(file2.filename)
            file_path2 = os.path.join(app.config['UPLOAD_FOLDER'], filename2)
            file2.save(file_path2)

            # Load data from files
            data1 = load_data(file_path1)
            data2 = load_data(file_path2)

            # Perform validations
            record_count_result = record_count_check(data1, data2)
            random_check_result = check_random_record(data1, data2, column_name1, column_name2)
            matched_columns, unmatched_columns = compare_column_records(data1, data2)

            print("\n*-------------------------------------------------------------*\n")
            print("record_count_result - {record_count_result}")
            print(record_count_result)
            print("\n*-------------------------------------------------------------*\n")
            print("random_check_result - {random_check_result}")
            print(random_check_result)
            print("\n*-------------------------------------------------------------*\n")
            print("matched_columns - {matched_columns}")
            print(matched_columns)
            print("\n*-------------------------------------------------------------*\n")
            print("unmatched_columns - {unmatched_columns}")
            print(unmatched_columns)
            print("\n*-------------------------------------------------------------*\n")
            if random_check_result['unmatched_columns']:
                print("\n*-------------------------------------------------------------*\n")
                print("Unmatched Data:")
                for column in random_check_result['unmatched_columns']:
                    value1 = random_check_result['random_record_data'].get(column, "N/A")
                    value2 = random_check_result['matching_record_data'].get(column, "N/A")
                    print(f"{column} - File 1: {value1} | File 2: {value2}")

            # Clean up temporary files
            os.remove(file_path1)
            os.remove(file_path2)

            # Render results
            return render_template('validate_data_form.html', record_count_result=record_count_result, random_check_result=random_check_result, matched_columns=matched_columns, unmatched_columns=unmatched_columns)
        else:
            return "Invalid file format or file not provided."
    return render_template('data_validation.html')


def check_ozbee_columns(folder_path):
    results = []
    for filename in os.listdir(folder_path):
        print(f"Processing file: {filename}")
        if filename.endswith(".csv") or filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            try:
                if filename.endswith(".csv"):
                    df = pd.read_csv(file_path, encoding='ISO-8859-1')
                else:
                    df = pd.read_excel(file_path)
            except Exception as e:
                print(f"Error reading file: {filename}, Error: {e}")
                results.append({
                    'filename': filename,
                    'has_ozbee_column': 'Error',
                    'contains_data': 'Error',
                    'error': f"Failed to read file: {e}"
                })
                continue

            # Print all columns in the file
            print(f"Columns in {filename}: {list(df.columns)}")

            # Check for 'ozbee__' columns in a case-insensitive manner
            ozbee_columns = [col for col in df.columns if col.lower().startswith('ozbee__')]
            print(f"Found ozbee columns in {filename}: {ozbee_columns}")
            if ozbee_columns:
                contains_data = any(df[col].notna().any() for col in ozbee_columns)
                results.append({
                    'filename': filename,
                    'has_ozbee_column': 'Yes',
                    'contains_data': 'Yes' if contains_data else 'No',
                    'error': None
                })
                print(f"File {filename} has ozbee columns with data: {contains_data}")
            else:
                results.append({
                    'filename': filename,
                    'has_ozbee_column': 'No',
                    'contains_data': 'No',
                    'error': None
                })
                print(f"File {filename} does not have ozbee columns")

    return results

@app.route('/check_ozbee_columns', methods=['GET', 'POST'])
def check_ozbee_columns_route():
    if request.method == 'POST':
        folder_path = request.form['folder_path']
        try:
            results = check_ozbee_columns(folder_path)
            return render_template('ozbee_columns_results.html', results=results)
        except Exception as e:
            return f"An error occurred: {str(e)}"
    else:
        return render_template('ozbee_columns_form.html')

if __name__ == '__main__':
    app.run(debug=True)
